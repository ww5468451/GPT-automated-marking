import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
import random
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
import selenium
import re
import openpyxl
import xlsxwriter

chrome_options = webdriver.ChromeOptions()
# 使用headless浏览器模式，这样就可以在不打开浏览器的情况下模拟操作
# chrome_options.add_argument('--headless')
# 使用无头模式下缺少浏览器信息，或默认填充的浏览器信息带有爬虫痕迹，所以给其加上请求头
# prefs = {"profile.managed_default_content_settings.images": 2, 'permissions.default.stylesheet': 2}
# chrome_options.add_experimental_option("prefs", prefs)  # 禁止加载图片和CSS
chrome_options.add_experimental_option("excludeSwitches", ['enable-automation'])
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('useAutomationExtension', False)
chrome_options.add_argument("--disable-javascript")  # 禁用JavaScript
# chrome_options.add_argument('blink-settings=imagesEnabled=false')
# chrome_options.page_load_strategy = 'eager'
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_argument(r'--user-data-dir=C:\Users\admin\AppData\Local\Google\Chrome\User Data3')
chrome_options.add_argument('--profile-directory=Default')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('window-size=1920x1080')  # 页面动态加载的时候，无头模式默认size为0x0
chrome_options.add_argument('--start-maximized')


driver1 = webdriver.Chrome(options=chrome_options)


def main(url):
    print('该产品的url为：' + str(url))
    # url = "https://www.amazon.fr/dp/B000ZSW6EK"
    try:
        driver1.get(url)
    except:
        time.sleep(2.0)
        driver1.refresh()
        driver1.get(url)
    m = 0
    title_text = ''
    time.sleep(8.0)
    while m < 2:
        try:
            title_text = driver1.find_element(by=By.ID, value='titleSection').text
            break
        except:
            time.sleep(20)
            driver1.refresh()
            m = m + 1
    try:
        price = driver1.find_element(by=By.ID, value='corePriceDisplay_desktop_feature_div').find_element(by=By.CLASS_NAME, value='a-price').text.replace('\n', '.').replace('SAR', '')
    except:
        price = " "
    def have_product():
        try:
            time.sleep(3.0)
            see_more = driver1.find_element(by=By.LINK_TEXT, value='See more')
            see_more.click()
        except:
            pass
        time.sleep(3.0)
        try:
            close_= driver1.find_element(by=By.XPATH, value='//*[@id="a-popover-2"]/div/header/button')
            close_.click()
        except:
            pass
        time.sleep(3.0)

        try:
            time.sleep(2.0)
            shuoming = driver1.find_element(by=By.ID, value='glance_icons_div').text
        except:
            shuoming = ''

        try:
            products = driver1.find_element(by=By.ID, value='productOverview_feature_div').text
        except:
            products = ''
        # print(products)
        try:
            five_dec = driver1.find_element(by=By.ID, value='feature-bullets').text
        except:
            five_dec = ''
        # try:
        #     page_text = driver1.find_element(by=By.ID, value="aplusBatch_feature_div").text
        # except:
        #     page_text = ''
        try:
            expan = driver1.find_element(by=By.XPATH, value='//*[@id="voyager-expand-all-btn"]')
            expan.click()
        except:
            pass

        time.sleep(5.0)
        try:
            product_description = driver1.find_element(by=By.ID, value='prodDetails').text
        except selenium.common.exceptions.NoSuchElementException:
            product_description = ''

        try:
            product_ = driver1.find_element(by=By.ID, value='productDetails_feature_div').text
        except:
            product_ = ''

        try:
            miaoshu = driver1.find_element(by=By.ID, value='productDescription').text
        except:
            miaoshu = ''
        print(miaoshu)

        product_descriptions = products + five_dec + product_description + product_ + miaoshu
        product_descriptions = product_descriptions.replace('Materials', 'Material')
        # print(product_descriptions)
        # print(product_descriptions)
        # 下拉滚动条
        current_scroll_position = 2500
        driver1.execute_script("window.scrollTo(0, {});".format(current_scroll_position))
        try:
            expand = driver1.find_element(by=By.ID, value='voyager-expand-all-btn')
            expand.click()
        except:
            pass
        time.sleep(2.0)
        expand = ''
        try:
            expand = driver1.find_element(by=By.ID, value='productDetails_expanderSectionTables').get_attribute('innerHTML').replace('\n', '').replace('&amp;', '&')
            expand = re.sub('>\s+<', '><', expand)
            expand = re.sub('>\s+', '>', expand)
            expand = re.sub('\s+<', '<', expand)
            # print(expand)
        except:
            pass

        if product_descriptions != '':
            try:
                dimensions = re.search(r'Product Dimensions\s*(.*?)(?:\n|$)', product_descriptions.replace('Materials & Care', '')).group(1)

            except:
                try:
                    dimensions = re.search(r'Package Dimensions\s*(.*?)(?:\n|$)', product_descriptions.replace('Materials & Care', '')).group(1)
                except:
                    dimensions = " "
            try:
                color = re.search(r'Color\s*(.*?)(?:\n|$)', product_descriptions).group(1)

            except:

                color = " "
            try:
                style = re.search(r'Style\s*(.*?)(?:\n|$)', product_descriptions).group(1)

            except:

                style = " "
            try:
                shape = re.search(r'Shape\s*(.*?)(?:\n|$)', product_descriptions).group(1)

            except:

                shape = " "
            shape = shape.replace('About this item','')


            print('来到这一步')
            try:
                materials = []
                miaoshu = product_descriptions.replace('Materials & Care', '').replace('\n', '换行').split(
                    '换行')
                # print(product_descriptions)
                for j in miaoshu:
                    materials_ = re.findall(r'(.*?)Material\s*(.*?)(?:\n|$)', j, re.DOTALL)
                    # print(materials_)
                    for i in materials_:
                        i = [m for m in i]
                        print(i)
                        if i[0] == '':
                            i = [i[1].replace(': ', '')]
                        material = ''.join(i)
                        materials.append(material)
                material = ';'.join(materials).replace('Material Type', 'Material ').replace('Material', 'Material ')
            except:
                material = ''
            if material == '' or 'Type' in material:

                try:
                    material = re.findall(r'>Materials & Care</span>(.*?)</tbody>', expand, re.DOTALL)[0]
                    material = re.sub(r'<[^>]+>', '', material)
                    material = material.replace('Material Type','Material ').replace('Material','Material ')
                except:
                    material = ''

            material = material.replace('s: ', '')
            # print(dimensions+"#"+material)
            print("#######################################################################################################")

            return dimensions+"#"+material+"#"+color+"#"+style+"#"+shape,  dimensions, material, color, style, shape
        else:

            return " "

    if title_text != '':
        random_num = round(random.uniform(8, 10), 2)
        time.sleep(random_num)
        title = driver1.find_element(by=By.ID, value='titleSection')
        title.click()
        # text = title_text+"#"+price
        text = have_product()
        return text
    else:
        print('该ASIN在亚马逊上已不存在')
        return ''


def get_url(country):
    ASIN0 = open('ASIN.txt', 'r+', encoding='utf-8')
    ASINS = ASIN0.readlines()
    if country == '美国':
        urls = ['https://www.amazon.com/dp/' + str(i).replace('\n', '') for i in ASINS]
    elif country == '德国':
        urls = ['https://www.amazon.de/dp/' + str(i).replace('\n', '') for i in ASINS]
    elif country == '英国':
        urls = ['https://www.amazon.co.uk/dp/' + str(i).replace('\n', '') for i in ASINS]
    elif country == '法国':
        urls = ['https://www.amazon.fr/dp/' + str(i).replace('\n', '') for i in ASINS]
    elif country == '意大利':
        urls = ['https://www.amazon.it/dp/' + str(i).replace('\n', '') for i in ASINS]
    elif country == '西班牙':
        urls = ['https://www.amazon.es/dp/' + str(i).replace('\n', '') for i in ASINS]
    else:
        urls = ['https://www.amazon.com/dp/' + str(i).replace('\n', '') for i in ASINS]
    return urls


if __name__ == '__main__':
    country = '美国'
    urls = get_url(country)
    m = 1
    path = input('请输入抓取数据的保存路径,如(D:\科研ting\数据\打标数据\材质-尺寸.xlsx)')
    try:
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook['Sheet1']
        workbook.save(path)
        workbook.close()
    except FileNotFoundError:
        workbook = xlsxwriter.Workbook(path)
        worksheet = workbook.add_worksheet()
        workbook.close()

    workbook = openpyxl.load_workbook(path)
    worksheet_1 = workbook['Sheet1']
    worksheet_1['A1'] = 'ASIN'
    worksheet_1['B1'] = '材质'
    worksheet_1['C1'] = '尺寸'
    worksheet_1['D1'] = '颜色'
    worksheet_1['E1'] = '风格'
    worksheet_1['F1'] = '形状'
    for url in urls:
        ASIN = url.split('/')[-1]
        print(ASIN)
        print(f'*******************************************这是第{m}个产品************************************************')
        text, dimensions, material, color, style, shape = main(url)
        random_num = round(random.uniform(2, 3), 2)
        time.sleep(random_num)
        if text != '':
            try:
                print(ASIN + "#" + text)
                row = worksheet_1.max_row + 1
                worksheet_1['A' + str(row)] = ASIN
                worksheet_1['B' + str(row)] = material
                worksheet_1['C' + str(row)] = dimensions
                worksheet_1['D' + str(row)] = color
                worksheet_1['E' + str(row)] = style
                worksheet_1['F' + str(row)] = shape
                workbook.save(path)
                random_num = round(random.uniform(5, 12), 2)
                time.sleep(random_num)

            except TypeError:
                pass
        else:
            pass

        m = m + 1
    driver1.close()

