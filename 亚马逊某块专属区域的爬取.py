import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
import random
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
import selenium
import re
import openpyxl
import xlsxwriter

chrome_options = webdriver.ChromeOptions()
# 使用headless浏览器模式，这样就可以在不打开浏览器的情况下模拟操作
# chrome_options.add_argument('--headless')
# 使用无头模式下缺少浏览器信息，或默认填充的浏览器信息带有爬虫痕迹，所以给其加上请求头
prefs = {"profile.managed_default_content_settings.images": 2, 'permissions.default.stylesheet': 2}
chrome_options.add_experimental_option("prefs", prefs)  # 禁止加载图片和CSS
chrome_options.add_experimental_option("excludeSwitches", ['enable-automation'])
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('useAutomationExtension', False)
chrome_options.add_argument("--disable-javascript") #禁用JavaScript
# chrome_options.add_argument('blink-settings=imagesEnabled=false')
# chrome_options.page_load_strategy = 'eager'
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_argument(r'--user-data-dir=C:\Users\Administrator\AppData\Local\Google\Chrome\User Data')
chrome_options.add_argument('--profile-directory=Default')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('window-size=1920x1080')# 页面动态加载的时候，无头模式默认size为0x0
chrome_options.add_argument('--start-maximized')
# chrome_options.add_argument('Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Mobile Safari/537.36')

driver1 = webdriver.Chrome(options=chrome_options)
def main(url):
    print('该产品的url为：' + str(url))
    # url = "https://www.amazon.fr/dp/B000ZSW6EK"
    try:
        driver1.get(url)
    except:
        time.sleep(2.0)
        driver1.refresh()
        driver1.get(url)
    m = 0
    title_text = ''
    while m < 5:
        try:
            title_text = driver1.find_element(by=By.ID, value='titleSection').text
            break
        except:
            time.sleep(20)
            driver1.refresh()
            m = m + 1

    def have_product():
        i = 0
        while i < 3:
            try:
                time.sleep(3.0)
                table = driver1.find_element(by=By.ID, value="glance_icons_div").text
                if table !="":
                    try:
                        material = re.search(r'Material\s*(.*?)(?:\n|$)', table).group(1)
                    except:
                        material = " "

                    try:
                        mounting = re.search(r'Mounting Type\s*(.*?)(?:\n|$)', table).group(1)
                    except:
                        mounting = " "

                    try:
                        shelf_type = re.search(r'Shelf Type\s*(.*?)(?:\n|$)', table).group(1)
                    except:
                        shelf_type = " "

                    try:
                        shape = re.search(r'Shape\s*(.*?)(?:\n|$)', table).group(1)
                    except:
                        shape = " "

                    return material+"#"+mounting+"#"+shelf_type+"#"+shape
            except:
                time.sleep(6.0)
                i = i + 1
                driver1.refresh()
    if title_text != '':
        random_num = round(random.uniform(8, 10), 2)
        time.sleep(random_num)
        title = driver1.find_element(by=By.ID, value='titleSection')
        title.click()
        text = have_product()
        return text
    else:
        print('该ASIN在亚马逊上已不存在')
        return ''

def get_url():
    ASIN0 = open('ASIN.txt', 'r+', encoding='utf-8')
    ASINS = ASIN0.readlines()
    urls = ['https://www.amazon.com/dp/'+str(i).replace('\n', '') for i in ASINS]
    return urls

if __name__ == '__main__':
    urls = get_url()
    m = 1
    # 必须要搞一个10次左右就重新发起一个对话，不然就卡住动不了了，刷新也没有用
    try:
        workbook = openpyxl.load_workbook(r'D:\科研ting\数据\打标数据\材质-尺寸.xlsx')
        worksheet = workbook['Sheet1']
        workbook.save(r'D:\科研ting\数据\打标数据\材质-尺寸.xlsx')
        workbook.close()
    except FileNotFoundError:
        workbook = xlsxwriter.Workbook(r'D:\科研ting\数据\打标数据\材质-尺寸.xlsx')
        worksheet = workbook.add_worksheet()
        workbook.close()

    workbook = openpyxl.load_workbook(r'D:\科研ting\数据\打标数据\材质-尺寸.xlsx')
    worksheet_1 = workbook['Sheet1']
    worksheet_1['A1'] = 'ASIN'
    worksheet_1['B1'] = '材质'
    worksheet_1['C1'] = '尺寸'

    for url in urls:
        ASIN = url.split('/')[-1]
        print(ASIN)
        print(f'*******************************************这是第{m}个产品************************************************')
        text, dimensions, material = main(url)
        random_num = round(random.uniform(2, 3), 2)
        time.sleep(random_num)
        if text != '':
            try:
                print(ASIN + "#" + text)
                row = worksheet_1.max_row + 1
                worksheet_1['A' + str(row)] = ASIN
                worksheet_1['B' + str(row)] = material
                worksheet_1['C' + str(row)] = dimensions
                random_num = round(random.uniform(5, 12), 2)
                time.sleep(random_num)

            except TypeError:
                pass
        else:
            pass

        m = m + 1
    driver1.close()
